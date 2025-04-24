import sys
from openpyxl import load_workbook
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print("Please provide the Excel filename as an argument.")
    sys.exit(1)

filename = sys.argv[1]

# Load the workbook
workbook = load_workbook(filename=filename)

# Loop through all sheets and all cells to make the content bold
for sheet in workbook:
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True)

# Save the modified workbook with "_modified" appended to the original filename
modified_filename = filename.replace(".xlsx", "_modified.xlsx")
workbook.save(filename=modified_filename)
